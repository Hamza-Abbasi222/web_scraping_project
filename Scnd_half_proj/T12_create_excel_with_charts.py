# T12_create_excel_with_charts.py

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Create a new workbook
wb = Workbook()
ws = wb.active

# Sample data (replace this with your actual data)
data = [
    ("Product", "Sales"),
    ("Product A", 1000),
    ("Product B", 1500),
    ("Product C", 800),
    ("Product D", 2000)
]

# Write data to worksheet
for row in data:
    ws.append(row)

# Create a bar chart
bar_chart = BarChart()
bar_chart.title = "Sales by Product"
bar_chart.y_axis.title = "Sales"
bar_chart.x_axis.title = "Product"

# Define data for the chart
values = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(data))
categories = Reference(ws, min_col=1, min_row=2, max_row=len(data))

# Add data to the chart
bar_chart.add_data(values, titles_from_data=True)
bar_chart.set_categories(categories)

# Add the chart to the worksheet
ws.add_chart(bar_chart, "E1")

# Add a formula to calculate total sales
ws.cell(row=len(data) + 2, column=1, value="Total Sales")
ws.cell(row=len(data) + 2, column=2, value=f"=SUM(B2:B{len(data) + 1})")

# Save the workbook
wb.save("sales_data_with_charts.xlsx")
