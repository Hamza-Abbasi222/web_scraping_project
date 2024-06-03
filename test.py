# T16_Project.py

import requests
from bs4 import BeautifulSoup
import pandas as pd
from fpdf import FPDF
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import re

# Define the URL
base_url = "http://books.toscrape.com/catalogue/page-{}.html"

# Function to fetch data from the website
def fetch_data():
    all_books = []
    for page in range(1, 3):  # Adjust the range as needed
        url = base_url.format(page)
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        books = soup.find_all('article', class_='product_pod')
        
        for book in books:
            title = book.h3.a['title']
            price = book.find('p', class_='price_color').text
            availability = book.find('p', class_='instock availability').text.strip()
            all_books.append({'Title': title, 'Price': price, 'Availability': availability})
    
    return pd.DataFrame(all_books)

# Function to process and clean data
def process_data(df):
    df['Price'] = df['Price'].apply(lambda x: re.sub(r'[^0-9.]', '', x)).astype(float)
    return df

# Function to generate reports
def generate_reports(df):
    # Export to CSV
    csv_file = f"book_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    df.to_csv(csv_file, index=False)
    
    # Export to Excel
    excel_file = f"book_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    df.to_excel(excel_file, index=False, sheet_name='Books')

    # Export to PDF
    pdf_file = f"book_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    for i in range(len(df)):
        pdf.cell(200, 10, txt=f"Title: {df.iloc[i]['Title']}", ln=True)
        pdf.cell(200, 10, txt=f"Price: £{df.iloc[i]['Price']}", ln=True)
        pdf.cell(200, 10, txt=f"Availability: {df.iloc[i]['Availability']}", ln=True)
        pdf.cell(200, 10, txt="-"*50, ln=True)
    
    pdf.output(pdf_file)

    return csv_file, excel_file, pdf_file

# Function to create a pivot table and export it to Excel
def create_pivot_table(df):
    pivot_df = df.pivot_table(index='Availability', values='Price', aggfunc='mean').reset_index()
    pivot_file = f"pivot_table_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    pivot_df.to_excel(pivot_file, index=False, sheet_name='Pivot Table')

    return pivot_file

# Function to create a bar chart for sales by product line
def create_bar_chart(df):
    plt.figure(figsize=(10, 6))
    df['Price'].value_counts().plot(kind='bar')
    plt.title('Price Distribution of Books')
    plt.xlabel('Price')
    plt.ylabel('Number of Books')
    plt.savefig('bar_chart.png')

# Main function to run the tasks
def main():
    # Fetch data
    df = fetch_data()
    print("Data fetched successfully.")
    
    # Process and clean data
    df = process_data(df)
    print("Data processed and cleaned successfully.")
    
    # Generate reports
    csv_file, excel_file, pdf_file = generate_reports(df)
    print(f"Reports generated and saved as {csv_file}, {excel_file}, {pdf_file}.")
    
    # Create pivot table and export to Excel
    pivot_file = create_pivot_table(df)
    print(f"Pivot table created and saved as {pivot_file}.")
    
    # Create bar chart
    create_bar_chart(df)
    print("Bar chart created and saved as bar_chart.png.")

# Execute the main function
if __name__ == "__main__":
    main()
